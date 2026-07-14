import streamlit as st
import pandas as pd
import numpy as np
import math
import io
import openpyxl
from openpyxl.styles import Font, Alignment
import datetime
from streamlit_cookies_manager import EncryptedCookieManager
import os
import altair as alt

# ===================== 用户数据库 =====================
USER_CREDENTIALS = {
    "msj01": "888888",
    "cyt01": "888888",
    "user01": "000000"
}

# ===================== 0. 登录与 Cookie =====================
cookies = EncryptedCookieManager(
    password=os.environ.get("COOKIES_PASSWORD", "a_very_secret_password_12345")
)
if not cookies.ready():
    st.stop()

def check_login():
    if cookies.get("auth_status") == "logged_in":
        return True
    st.title("能源系统分析 - 身份验证")
    with st.form("login_form"):
        user_input = st.text_input("账号")
        pw_input = st.text_input("密码", type="password")
        submit = st.form_submit_button("登录")
        if submit:
            if user_input in USER_CREDENTIALS and USER_CREDENTIALS[user_input] == pw_input:
                cookies["auth_status"] = "logged_in"
                cookies["current_user"] = user_input
                cookies.save()
                st.success(f"欢迎回来，{user_input}！正在进入系统...")
                st.rerun()
                return True
            else:
                st.error("账号或密码不正确，请重试")
    return False

def logout():
    current_user = cookies.get("current_user", "未知用户")
    st.sidebar.write(f"当前用户: **{current_user}**")
    if st.sidebar.button("退出登录"):
        cookies["auth_status"] = "logged_out"
        cookies["current_user"] = ""
        cookies.save()
        st.rerun()
    # ===== 新增：8760 数据模板下载 =====
    st.sidebar.markdown("---")
    st.sidebar.markdown("**数据模板下载**")
    template_csv = "PV_Unit_Output(kWh),Wind_Unit_Output(kWh),Load(kWh)\n"
    st.sidebar.download_button(
        label="下载逐时数据模板 (CSV)",
        data=template_csv,
        file_name="8760_数据模板.csv",
        mime="text/csv",
        use_container_width=True,
    )
    st.sidebar.caption("上传前请确保文件包含 8760 行逐时数据。")
    
# ===================== 1. 工具函数 =====================
def generate_8760_month_array():
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    hours_in_month = [d * 24 for d in days_in_month]
    month_array = np.zeros(8760, dtype=int)
    current_hour_index = 0
    for month in range(1, 13):
        num_hours = hours_in_month[month - 1]
        if current_hour_index + num_hours > 8760:
            num_hours = 8760 - current_hour_index
        month_array[current_hour_index:current_hour_index + num_hours] = month
        current_hour_index += num_hours
    return month_array

# ===================== 2. 核心计算 =====================
def calculate_single_case(
    pv_unit_data, wind_unit_data, load_data,
    pv_capacity_1mw, wind_capacity_1mw,
    storage_power_mw, storage_duration_h,
    storage_efficiency_p3, discharge_depth_p2,
    peak_valley_map, prices, month_8760_array,
    discharge_allowed, return_hourly=False,
    hourly_self_prices=None, hourly_on_grid_prices=None
):
    pv_generation = pv_unit_data * pv_capacity_1mw
    wind_generation = wind_unit_data * wind_capacity_1mw
    generation_data = pv_generation + wind_generation

    storage_capacity_max_kwh = storage_power_mw * storage_duration_h * 1000
    max_charge_power_kwh = storage_power_mw * 1000
    max_discharge_power_kwh = storage_power_mw * 1000
    storage_efficiency_single = math.sqrt(storage_efficiency_p3)
    storage_capacity_arr = np.zeros(8760)

    total_consumption_sum = 0.0
    total_on_grid_sum = 0.0
    total_charge_loss = 0.0
    total_discharge_loss = 0.0
    total_discharge_energy = 0.0

    time_period_stats = {
        '尖峰': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '峰':   {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '平':   {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '谷':   {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0},
        '深谷': {'consumption': 0.0, 'on_grid': 0.0, 'consumption_cost_sum': 0.0, 'on_grid_cost_sum': 0.0}
    }

    if return_hourly:
        hourly_pv_gen = np.zeros(8760)
        hourly_wind_gen = np.zeros(8760)
        hourly_generation = np.zeros(8760)
        hourly_load = np.zeros(8760)
        hourly_charge_in = np.zeros(8760)
        hourly_charge_source = np.zeros(8760)
        hourly_discharge_req = np.zeros(8760)
        hourly_discharge_out = np.zeros(8760)
        hourly_soc = np.zeros(8760)
        hourly_consumption = np.zeros(8760)
        hourly_on_grid = np.zeros(8760)
        hourly_period = ['平'] * 8760
        hourly_charge_loss = np.zeros(8760)
        hourly_discharge_loss = np.zeros(8760)

    use_hourly_prices = (hourly_self_prices is not None) and (hourly_on_grid_prices is not None)

    for i in range(8760):
        current_generation = generation_data[i]
        current_load = load_data[i]
        previous_storage_capacity = storage_capacity_arr[i - 1] if i > 0 else 0.0
        hour_of_day = i % 24
        month = month_8760_array[i]
        current_period_type = peak_valley_map.get(f"{hour_of_day}_{month}", '平')

        # 充电
        storage_charge_in_i = 0.0
        storage_charge_source = 0.0
        on_grid_i = 0.0
        if current_generation > current_load:
            available_from_generation = current_generation - current_load
            remaining_storage_capacity = storage_capacity_max_kwh - previous_storage_capacity
            charge_effective = available_from_generation * storage_efficiency_single
            max_storage_charge = max_charge_power_kwh * storage_efficiency_single
            storage_charge_in_i = min(charge_effective, remaining_storage_capacity,
                                      max_storage_charge / storage_efficiency_single)
            storage_charge_in_i = max(0, storage_charge_in_i)
            storage_charge_source = storage_charge_in_i / storage_efficiency_single
            total_charge_loss += (storage_charge_source - storage_charge_in_i)
            on_grid_i = max(0, available_from_generation - storage_charge_source)

        # 放电（自动检测小时制/时段制）
        storage_required_discharge_i = 0.0
        storage_discharge_out_i = 0.0
        _first_dk = next(iter(discharge_allowed), None) if discharge_allowed else None
        _is_hourly = isinstance(_first_dk, int) if _first_dk is not None else False
        allow_discharge = discharge_allowed.get(hour_of_day, False) if _is_hourly else discharge_allowed.get(current_period_type, False)
        if current_load > current_generation and allow_discharge:
            load_gap = current_load - current_generation
            min_storage_capacity = storage_capacity_max_kwh * (1 - discharge_depth_p2)
            max_discharge_from_current_storage = max(0, previous_storage_capacity - min_storage_capacity)
            required_discharge_for_load = load_gap / storage_efficiency_single
            storage_required_discharge_i = min(max_discharge_from_current_storage,
                                               required_discharge_for_load,
                                               max_discharge_power_kwh)
            storage_required_discharge_i = max(0, storage_required_discharge_i)
            storage_discharge_out_i = storage_required_discharge_i * storage_efficiency_single
            total_discharge_loss += (storage_required_discharge_i - storage_discharge_out_i)
            total_discharge_energy += storage_required_discharge_i

        storage_capacity_i = previous_storage_capacity + storage_charge_in_i - storage_required_discharge_i
        storage_capacity_arr[i] = max(0, min(storage_capacity_i, storage_capacity_max_kwh))

        total_available = current_generation + storage_discharge_out_i
        consumption_i = min(total_available, current_load)
        total_consumption_sum += consumption_i

        if current_generation <= current_load and total_available > current_load:
            on_grid_i += (total_available - current_load)
        total_on_grid_sum += on_grid_i

        if return_hourly:
            hourly_pv_gen[i] = pv_generation[i]
            hourly_wind_gen[i] = wind_generation[i]
            hourly_generation[i] = current_generation
            hourly_load[i] = current_load
            hourly_charge_in[i] = storage_charge_in_i
            hourly_charge_source[i] = storage_charge_source
            hourly_discharge_req[i] = storage_required_discharge_i
            hourly_discharge_out[i] = storage_discharge_out_i
            hourly_soc[i] = storage_capacity_arr[i]
            hourly_consumption[i] = consumption_i
            hourly_on_grid[i] = on_grid_i
            hourly_period[i] = current_period_type
            hourly_charge_loss[i] = storage_charge_source - storage_charge_in_i
            hourly_discharge_loss[i] = storage_required_discharge_i - storage_discharge_out_i

        if use_hourly_prices:
            self_price_i = float(hourly_self_prices[i])
            on_grid_price_i = float(hourly_on_grid_prices[i])
        else:
            self_price_i = prices[current_period_type]['self']
            on_grid_price_i = prices[current_period_type]['on_grid']

        if current_period_type in time_period_stats:
            s = time_period_stats[current_period_type]
            s['consumption'] += consumption_i
            s['on_grid'] += on_grid_i
            s['consumption_cost_sum'] += consumption_i * self_price_i
            s['on_grid_cost_sum'] += on_grid_i * on_grid_price_i

    total_pv_gen = np.sum(pv_generation)
    total_wind_gen = np.sum(wind_generation)
    total_curtailment_sum = total_charge_loss + total_discharge_loss + storage_capacity_arr[-1]
    total_generation_sum = total_pv_gen + total_wind_gen
    total_consumption_cost = sum(s['consumption_cost_sum'] for s in time_period_stats.values())
    total_on_grid_cost = sum(s['on_grid_cost_sum'] for s in time_period_stats.values())
    weighted_self_price = total_consumption_cost / total_consumption_sum if total_consumption_sum > 0 else 0.0
    weighted_on_grid_price = total_on_grid_cost / total_on_grid_sum if total_on_grid_sum > 0 else 0.0
    total_revenue = total_consumption_cost + total_on_grid_cost + (total_curtailment_sum * prices['Curtailment'])
    integrated_price = total_revenue / total_generation_sum if total_generation_sum > 0 else 0.0
    equivalent_cycles = total_discharge_energy / storage_capacity_max_kwh if storage_capacity_max_kwh > 0 else 0.0
    pv_hours = total_pv_gen / (pv_capacity_1mw * 1000) if pv_capacity_1mw > 0 else 0.0
    wind_hours = total_wind_gen / (wind_capacity_1mw * 1000) if wind_capacity_1mw > 0 else 0.0

    result = {
        "total_generation_sum": total_generation_sum,
        "total_pv_generation": total_pv_gen,
        "total_wind_generation": total_wind_gen,
        "pv_hours": pv_hours,
        "wind_hours": wind_hours,
        "total_consumption_sum": total_consumption_sum,
        "total_on_grid_sum": total_on_grid_sum,
        "total_curtailment_sum": total_curtailment_sum,
        "weighted_self_price": weighted_self_price,
        "weighted_on_grid_price": weighted_on_grid_price,
        "integrated_price": integrated_price,
        "storage_equivalent_cycles": equivalent_cycles,
        "time_period_stats": time_period_stats
    }
    if return_hourly:
        result["hourly_data"] = {
            "hour": np.arange(8760),
            "month": month_8760_array,
            "hour_of_day": np.arange(8760) % 24,
            "period_type": hourly_period,
            "pv_generation_kwh": hourly_pv_gen,
            "wind_generation_kwh": hourly_wind_gen,
            "total_generation_kwh": hourly_generation,
            "load_kwh": hourly_load,
            "charge_source_kwh": hourly_charge_source,
            "charge_into_storage_kwh": hourly_charge_in,
            "charge_loss_kwh": hourly_charge_loss,
            "discharge_required_kwh": hourly_discharge_req,
            "discharge_out_kwh": hourly_discharge_out,
            "discharge_loss_kwh": hourly_discharge_loss,
            "storage_soc_kwh": hourly_soc,
            "consumption_kwh": hourly_consumption,
            "on_grid_kwh": hourly_on_grid,
        }
    return result

def perform_batch_calculation(pv_unit_data, wind_unit_data, load_data, params, month_8760_array):
    batch_results = []
    total_load_sum = np.sum(load_data)
    for current_pv in params['pv_list']:
        for current_wind in params['wind_list']:
            for current_power in params['power_list']:
                for current_duration in params['duration_list']:
                    res = calculate_single_case(
                        pv_unit_data, wind_unit_data, load_data,
                        current_pv, current_wind, current_power, current_duration,
                        params['efficiency'], params['depth'],
                        params['peak_valley_map'], params['prices'],
                        month_8760_array, params['discharge_allowed'],
                        hourly_self_prices=params.get('hourly_self_prices'),
                        hourly_on_grid_prices=params.get('hourly_on_grid_prices')
                    )
                    batch_results.append({
                        "光伏容量 (MW)": current_pv,
                        "风电容量 (MW)": current_wind,
                        "光伏利用小时数 (h)": res["pv_hours"],
                        "风电利用小时数 (h)": res["wind_hours"],
                        "储能功率 (MW)": current_power,
                        "储能时长 (h)": current_duration,
                        "储能容量 (MWh)": current_power * current_duration,
                        "加权自用电价": res["weighted_self_price"],
                        "加权上网电价": res["weighted_on_grid_price"],
                        "综合电价": res["integrated_price"],
                        "总发电量 (kWh)": res["total_generation_sum"],
                        "总消纳电量 (kWh)": res["total_consumption_sum"],
                        "总上网电量 (kWh)": res["total_on_grid_sum"],
                        "总折损电量 (kWh)": res["total_curtailment_sum"],
                        "自用比例 (%)": (res["total_consumption_sum"] / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "绿电占用电比例 (%)": (res["total_consumption_sum"] / total_load_sum * 100) if total_load_sum > 0 else 0.0,
                        "尖峰消纳 (%)": (res["time_period_stats"]["尖峰"]["consumption"] / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "峰消纳 (%)":   (res["time_period_stats"]["峰"]["consumption"]   / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "平消纳 (%)":   (res["time_period_stats"]["平"]["consumption"]   / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "谷消纳 (%)":   (res["time_period_stats"]["谷"]["consumption"]   / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "深谷消纳 (%)": (res["time_period_stats"]["深谷"]["consumption"] / res["total_generation_sum"] * 100) if res["total_generation_sum"] > 0 else 0.0,
                        "储能等效循环次数": res["storage_equivalent_cycles"]
                    })
    return batch_results

# ===================== 3. Excel 导出 =====================
def write_batch_results_to_excel(results, params):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "批量计算汇总"
    headers = [
        "序号", "光伏容量 (MW)", "风电容量 (MW)", "光伏利用小时数 (h)", "风电利用小时数 (h)",
        "储能功率 (MW)", "储能时长 (h)", "储能容量 (MWh)",
        "加权自用电价", "加权上网电价", "综合电价",
        "总发电量 (kWh)", "消纳总电量 (kWh)", "上网总电量 (kWh)", "折损总电量 (kWh)",
        "自用比例 (%)", "绿电占用电比例 (%)",
        "储能等效循环次数"
    ]
    sheet.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
    for row_idx, result in enumerate(results, 2):
        sheet.cell(row=row_idx, column=1, value=row_idx - 1)
        sheet.cell(row=row_idx, column=2, value=result['光伏容量 (MW)'])
        sheet.cell(row=row_idx, column=3, value=result['风电容量 (MW)'])
        sheet.cell(row=row_idx, column=4, value=round(result['光伏利用小时数 (h)'], 3))
        sheet.cell(row=row_idx, column=5, value=round(result['风电利用小时数 (h)'], 3))
        sheet.cell(row=row_idx, column=6, value=result['储能功率 (MW)'])
        sheet.cell(row=row_idx, column=7, value=result['储能时长 (h)'])
        sheet.cell(row=row_idx, column=8, value=result['储能容量 (MWh)'])
        sheet.cell(row=row_idx, column=9, value=round(result['加权自用电价'], 4))
        sheet.cell(row=row_idx, column=10, value=round(result['加权上网电价'], 4))
        sheet.cell(row=row_idx, column=11, value=round(result['综合电价'], 4))
        sheet.cell(row=row_idx, column=12, value=round(result['总发电量 (kWh)'], 6))
        sheet.cell(row=row_idx, column=13, value=round(result['总消纳电量 (kWh)'], 6))
        sheet.cell(row=row_idx, column=14, value=round(result['总上网电量 (kWh)'], 6))
        sheet.cell(row=row_idx, column=15, value=round(result['总折损电量 (kWh)'], 6))
        sheet.cell(row=row_idx, column=16, value=round(result['自用比例 (%)'], 2))
        sheet.cell(row=row_idx, column=17, value=round(result['绿电占用电比例 (%)'], 2))
        sheet.cell(row=row_idx, column=18, value=round(result['储能等效循环次数'], 2))
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

def write_hourly_data_to_excel(hourly_data, scheme_info):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "8760逐时过程量"
    summary_headers = [
        "光伏容量 (MW)", "风电容量 (MW)", "储能功率 (MW)", "储能时长 (h)",
        "储能容量 (MWh)", "综合电价", "总发电量 (kWh)", "总消纳电量 (kWh)",
        "总上网电量 (kWh)", "总折损电量 (kWh)", "自用比例 (%)", "绿电占用电比例 (%)"
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    summary_values = [
        scheme_info.get('光伏容量 (MW)', ''), scheme_info.get('风电容量 (MW)', ''),
        scheme_info.get('储能功率 (MW)', ''), scheme_info.get('储能时长 (h)', ''),
        scheme_info.get('储能容量 (MWh)', ''), round(scheme_info.get('综合电价', 0), 4),
        round(scheme_info.get('总发电量 (kWh)', 0), 2), round(scheme_info.get('总消纳电量 (kWh)', 0), 2),
        round(scheme_info.get('总上网电量 (kWh)', 0), 2), round(scheme_info.get('总折损电量 (kWh)', 0), 2),
        round(scheme_info.get('自用比例 (%)', 0), 2), round(scheme_info.get('绿电占用电比例 (%)', 0), 2),
    ]
    for col_idx, val in enumerate(summary_values, 1):
        cell = sheet.cell(row=2, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="center")
    start_row = 4
    hourly_headers = [
        "小时序号", "月份", "日内小时", "时段类型",
        "光伏发电 (kWh)", "风电发电 (kWh)", "总发电量 (kWh)", "负载 (kWh)",
        "充电来源 (kWh)", "充入储能 (kWh)", "充电损失 (kWh)",
        "放电需求 (kWh)", "放电输出 (kWh)", "放电损失 (kWh)",
        "储能SOC (kWh)", "消纳量 (kWh)", "上网量 (kWh)"
    ]
    for col_idx, header in enumerate(hourly_headers, 1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
    n = len(hourly_data["hour"])
    for i in range(n):
        row = start_row + 1 + i
        sheet.cell(row=row, column=1, value=i + 1)
        sheet.cell(row=row, column=2, value=int(hourly_data["month"][i]))
        sheet.cell(row=row, column=3, value=int(hourly_data["hour_of_day"][i]))
        sheet.cell(row=row, column=4, value=str(hourly_data["period_type"][i]))
        sheet.cell(row=row, column=5, value=round(float(hourly_data["pv_generation_kwh"][i]), 4))
        sheet.cell(row=row, column=6, value=round(float(hourly_data["wind_generation_kwh"][i]), 4))
        sheet.cell(row=row, column=7, value=round(float(hourly_data["total_generation_kwh"][i]), 4))
        sheet.cell(row=row, column=8, value=round(float(hourly_data["load_kwh"][i]), 4))
        sheet.cell(row=row, column=9, value=round(float(hourly_data["charge_source_kwh"][i]), 4))
        sheet.cell(row=row, column=10, value=round(float(hourly_data["charge_into_storage_kwh"][i]), 4))
        sheet.cell(row=row, column=11, value=round(float(hourly_data["charge_loss_kwh"][i]), 4))
        sheet.cell(row=row, column=12, value=round(float(hourly_data["discharge_required_kwh"][i]), 4))
        sheet.cell(row=row, column=13, value=round(float(hourly_data["discharge_out_kwh"][i]), 4))
        sheet.cell(row=row, column=14, value=round(float(hourly_data["discharge_loss_kwh"][i]), 4))
        sheet.cell(row=row, column=15, value=round(float(hourly_data["storage_soc_kwh"][i]), 4))
        sheet.cell(row=row, column=16, value=round(float(hourly_data["consumption_kwh"][i]), 4))
        sheet.cell(row=row, column=17, value=round(float(hourly_data["on_grid_kwh"][i]), 4))
    sheet.freeze_panes = f"A{start_row + 1}"
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

# ===================== 4. 辅助函数与 UI 组件 =====================
def init_session_state():
    if 'monthly_config_data' not in st.session_state:
        st.session_state.monthly_config_data = {
            m: {'尖峰': '19-21', '峰': '12-14, 17-18, 22-23', '平': '7-11, 15-16', '谷': '0-5', '深谷': '6-6'}
            for m in range(1, 13)
        }
    if 'batch_results' not in st.session_state:
        st.session_state.batch_results = []
    if 'discharge_allowed' not in st.session_state:
        st.session_state.discharge_allowed = {'尖峰': True, '峰': True, '平': True, '谷': True, '深谷': True}
    if 'price_mode' not in st.session_state:
        st.session_state.price_mode = 'period'
    if 'hourly_self_prices' not in st.session_state:
        st.session_state.hourly_self_prices = None
    if 'hourly_on_grid_prices' not in st.session_state:
        st.session_state.hourly_on_grid_prices = None
    if 'csv_discharge_str' not in st.session_state:
        st.session_state.csv_discharge_str = '0-23'

def parse_time_slot_input(s):
    try:
        slots = []
        for part in s.replace(' ', '').split(','):
            if not part: continue
            start, end = map(int, part.split('-'))
            slots.append((start, end))
        return slots
    except:
        return None

def parse_batch_input(input_str):
    try:
        parts = [float(x.strip()) for x in input_str.split(',')]
        if len(parts) == 1:
            return [parts[0]]
        elif len(parts) == 3:
            start, end, step = parts
            if step <= 0: return [start]
            return np.arange(start, end + 1e-9, step).tolist()
        else:
            return None
    except:
        return None

def get_final_map():
    final_map = {}
    for month in range(1, 13):
        h_map = {h: '平' for h in range(24)}
        for period in ['深谷', '谷', '平', '峰', '尖峰']:
            slots = parse_time_slot_input(st.session_state.monthly_config_data[month][period])
            if slots:
                for s, e in slots:
                    hrs = range(s, e+1) if s <= e else list(range(s, 24)) + list(range(0, e+1))
                    for h in hrs:
                        h_map[h] = period
        for h, p in h_map.items():
            final_map[f"{h}_{month}"] = p
    return final_map

def color_time_periods(val):
    color_map = {
        '尖峰': 'background-color: #ff6347; color: white',
        '峰':   'background-color: #ffd700; color: black',
        '平':   'background-color: #90ee90; color: black',
        '谷':   'background-color: #add8e6; color: black',
        '深谷': 'background-color: #4682b4; color: white',
    }
    return color_map.get(val, '')

# ===================== 5. 主 UI =====================
def main():
    logout()
    init_session_state()
    tab1, tab2 = st.tabs(["计算与分析", "电价与时段配置"])

    # ========== tab2：电价与时段配置 ==========
    with tab2:
        st.subheader("1. 电价模式选择")
        price_mode = st.radio(
            "选择电价模式",
            options=['period', 'csv'],
            format_func=lambda x: '分时段电价（默认）' if x == 'period' else 'CSV 逐时电价',
            horizontal=True,
            key="price_mode_radio"
        )
        st.session_state.price_mode = price_mode
        prices = {'Curtailment': st.number_input("折损/弃电电价 (元/kWh)", value=0.0, format="%.4f")}
        periods = ['尖峰', '峰', '平', '谷', '深谷']

        # ---- CSV 模式 ----
        if price_mode == 'csv':
            st.info("请上传包含 **8760 行** 逐时电价数据的 CSV 文件，需包含两列：**自用电价** 和 **上网电价**（单位：元/kWh）")
            csv_price_file = st.file_uploader("上传逐时电价 CSV", type="csv", key="price_csv_uploader")
            if csv_price_file is not None:
                try:
                    price_df = pd.read_csv(csv_price_file)
                    if len(price_df) != 8760:
                        st.error(f"CSV 文件行数必须为 8760，当前为 {len(price_df)} 行。")
                        st.session_state.hourly_self_prices = None
                        st.session_state.hourly_on_grid_prices = None
                    else:
                        cols = price_df.columns.tolist()
                        st.success(f"已读取 {len(price_df)} 行数据，列名：{cols}")
                        c1, c2 = st.columns(2)
                        with c1:
                            self_col = st.selectbox("自用电价对应列", cols, index=0, key="self_col_select")
                        with c2:
                            on_grid_col = st.selectbox("上网电价对应列", cols, index=min(1, len(cols)-1), key="on_grid_col_select")
                        st.session_state.hourly_self_prices = price_df[self_col].values.astype(float)
                        st.session_state.hourly_on_grid_prices = price_df[on_grid_col].values.astype(float)
                        st.markdown("**逐时电价预览（前 24 小时）：**")
                        preview_df = pd.DataFrame({
                            "小时": range(1, 25),
                            "自用电价 (元/kWh)": st.session_state.hourly_self_prices[:24],
                            "上网电价 (元/kWh)": st.session_state.hourly_on_grid_prices[:24]
                        })
                        st.dataframe(preview_df, height=(len(preview_df)+1)*35+3, use_container_width=True)
                except Exception as e:
                    st.error(f"解析 CSV 文件失败：{e}")
                    st.session_state.hourly_self_prices = None
                    st.session_state.hourly_on_grid_prices = None
            else:
                st.session_state.hourly_self_prices = None
                st.session_state.hourly_on_grid_prices = None
            for p in periods:
                prices[p] = {'self': 0.0, 'on_grid': 0.0}

        # ---- 分时段模式 ----
        if price_mode == 'period':
            st.subheader("时段电价")
            header_cols = st.columns([0.8] + [1]*5)
            with header_cols[0]:
                st.markdown("**时段 →**")
            for i, p in enumerate(periods):
                with header_cols[i+1]:
                    st.markdown(f"**{p}**")
            row1_cols = st.columns([0.8] + [1]*5)
            with row1_cols[0]:
                st.markdown("**自用电价**")
            for i, p in enumerate(periods):
                with row1_cols[i+1]:
                    prices[p] = prices.get(p, {})
                    prices[p]['self'] = st.number_input(
                        label="", value=0.3 if p == '深谷' else (1.2 if i < 2 else 0.6),
                        format="%.4f", key=f"s_{p}", label_visibility="collapsed"
                    )
            row2_cols = st.columns([0.8] + [1]*5)
            with row2_cols[0]:
                st.markdown("**上网电价**")
            for i, p in enumerate(periods):
                with row2_cols[i+1]:
                    prices[p]['on_grid'] = st.number_input(
                        label="", value=0.38, format="%.4f", key=f"o_{p}", label_visibility="collapsed"
                    )

        # ---- 允许放电时段 ----
        st.markdown("---")
        st.subheader("2. 允许放电时段")
        if price_mode == 'csv':
            st.caption("输入允许放电的小时范围（0–23），多个时段用逗号分隔，适用于全年每日")
            csv_discharge_input = st.text_input(
                "允许放电时段",
                value=st.session_state.csv_discharge_str,
                placeholder="例：7-22 或 0-5, 19-23",
                key="csv_discharge_text"
            )
            st.session_state.csv_discharge_str = csv_discharge_input
            try:
                allowed_set = set()
                for part in csv_discharge_input.replace(' ', '').split(','):
                    if not part: continue
                    if '-' in part:
                        s, e = part.split('-')
                        for h in range(int(s), int(e)+1):
                            if 0 <= h <= 23: allowed_set.add(h)
                    else:
                        h = int(part)
                        if 0 <= h <= 23: allowed_set.add(h)
                st.session_state.discharge_allowed = {h: (h in allowed_set) for h in range(24)}
                st.caption(f"已解析：允许放电 {len(allowed_set)} 小时 → {sorted(allowed_set)}")
            except:
                st.warning("格式解析失败，请检查输入。当前默认全部允许。")
                st.session_state.discharge_allowed = {h: True for h in range(24)}
        else:
            if any(isinstance(k, int) for k in st.session_state.discharge_allowed.keys()):
                st.session_state.discharge_allowed = {'尖峰': True, '峰': True, '平': True, '谷': True, '深谷': True}
            row3_cols = st.columns([0.8] + [1]*5)
            with row3_cols[0]:
                st.markdown("**允许放电**")
            for i, p in enumerate(periods):
                with row3_cols[i+1]:
                    st.session_state.discharge_allowed[p] = st.checkbox(
                        label="", value=st.session_state.discharge_allowed.get(p, True),
                        key=f"d_{p}", label_visibility="collapsed"
                    )

        # ---- 月度时段详细配置（CSV 模式隐藏） ----
        if price_mode != 'csv':
            st.markdown("---")
            st.subheader("3. 月度时段详细配置")
            st.info("格式: 开始-结束 (0-23)，多个时段用逗号分隔。优先级：尖峰 > 峰 > 平 > 谷 > 深谷")
            edit_col1, _ = st.columns([1, 2])
            with edit_col1:
                month = st.selectbox("当前编辑月份", range(1, 13))
            m_cols = st.columns(5)
            temp_inputs = {}
            for i, p in enumerate(['尖峰', '峰', '平', '谷', '深谷']):
                temp_inputs[p] = m_cols[i].text_input(
                    f"{p}时段定义",
                    value=st.session_state.monthly_config_data[month][p],
                    key=f"input_{month}_{p}"
                )
                st.session_state.monthly_config_data[month][p] = temp_inputs[p]
            with st.expander("批量复制当前月份配置到其他月份"):
                target_months = st.multiselect("选择目标月份", [m for m in range(1, 13) if m != month])
                if st.button("执行批量同步"):
                    if target_months:
                        for tm in target_months:
                            st.session_state.monthly_config_data[tm] = temp_inputs.copy()
                        st.success(f"已成功将 {month} 月配置同步至 {target_months} 月")
                        st.rerun()
                    else:
                        st.warning("请先选择目标月份")

        # ---- 4. 全年日均逐时电价趋势（Altair 静态图） ----
        st.markdown("---")
        st.subheader("4. 全年日均逐时电价趋势")
        if (st.session_state.price_mode == 'csv'
                and st.session_state.hourly_self_prices is not None
                and st.session_state.hourly_on_grid_prices is not None):
            hourly_prices_self = np.zeros(24)
            hourly_prices_on_grid = np.zeros(24)
            for h in range(24):
                hourly_prices_self[h] = np.mean(st.session_state.hourly_self_prices[h::24])
                hourly_prices_on_grid[h] = np.mean(st.session_state.hourly_on_grid_prices[h::24])
        else:
            f_map = get_final_map()
            hourly_prices_self = np.zeros(24)
            hourly_prices_on_grid = np.zeros(24)
            for h in range(24):
                total_self = 0.0
                total_on_grid = 0.0
                for m in range(1, 13):
                    period = f_map.get(f"{h}_{m}", "平")
                    total_self += prices[period]['self']
                    total_on_grid += prices[period]['on_grid']
                hourly_prices_self[h] = total_self / 12
                hourly_prices_on_grid[h] = total_on_grid / 12

        chart_data = pd.DataFrame({
            "小时": [f"{h:02d}:00" for h in range(24)],
            "日均自用电价 (元/kWh)": hourly_prices_self,
            "日均上网电价 (元/kWh)": hourly_prices_on_grid
        })

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**日均自用电价趋势**")
            alt_self = alt.Chart(chart_data).mark_line(color="#ff6347").encode(
                x=alt.X("小时", sort=None),
                y=alt.Y("日均自用电价 (元/kWh)", scale=alt.Scale(padding=50))
            ).properties(height=300)
            st.altair_chart(alt_self, use_container_width=True)

        with c2:
            st.markdown("**日均上网电价趋势**")
            alt_ongrid = alt.Chart(chart_data).mark_line(color="#1f77b4").encode(
                x=alt.X("小时", sort=None),
                y=alt.Y("日均上网电价 (元/kWh)", scale=alt.Scale(padding=50))
            ).properties(height=300)
            st.altair_chart(alt_ongrid, use_container_width=True)

    # ========== tab1：计算与分析 ==========
    with tab1:
        st.title("新能源项目源荷储匹配分析平台")
        f = st.file_uploader("上传 8760 逐时数据 (CSV)", type="csv")
        if f:
            df = pd.read_csv(f)
            c1, c2 = st.columns(2)
            eff = c1.number_input("储能往返效率 (P3)", value=0.85, max_value=1.0)
            dep = c2.number_input("储能放电深度 (P2)", value=0.9, max_value=1.0)

            st.markdown("---")
            st.subheader("批量计算参数设置")
            st.caption("格式：单个数字 (例: 1) 或 范围步长 (例: 1,4,1 代表从1到4，步长1)")
            bc1, bc2, bc3, bc4 = st.columns(4)
            with bc1:
                pv_raw = st.text_input("光伏容量 (MW)", value="5")
            with bc2:
                wi_raw = st.text_input("风电容量 (MW)", value="2")
            with bc3:
                p_raw = st.text_input("储能功率 (MW)", value="2,4,1")
            with bc4:
                d_raw = st.text_input("储能时长 (h)", value="1,4,1")

            pv_list = parse_batch_input(pv_raw)
            wi_list = parse_batch_input(wi_raw)
            p_list = parse_batch_input(p_raw)
            d_list = parse_batch_input(d_raw)

            if st.button("开始执行模拟计算", type="primary"):
                if None in [pv_list, wi_list, p_list, d_list]:
                    st.error("输入格式有误，请检查是否使用了英文逗号且格式正确。")
                else:
                    total_scenarios = len(pv_list) * len(wi_list) * len(p_list) * len(d_list)
                    st.info(f"即将计算 {total_scenarios} 组方案...")
                    params = {
                        "pv_list": pv_list, "wind_list": wi_list,
                        "power_list": p_list, "duration_list": d_list,
                        "efficiency": eff, "depth": dep,
                        "peak_valley_map": get_final_map(),
                        "prices": prices,
                        "discharge_allowed": st.session_state.discharge_allowed,
                        "hourly_self_prices": st.session_state.hourly_self_prices,
                        "hourly_on_grid_prices": st.session_state.hourly_on_grid_prices
                    }
                    with st.spinner("正在进行计算..."):
                        st.session_state.batch_results = perform_batch_calculation(
                            df["PV_Unit_Output(kWh)"].values,
                            df["Wind_Unit_Output(kWh)"].values,
                            df["Load(kWh)"].values,
                            params, generate_8760_month_array()
                        )
                    st.session_state.last_params = params
                    st.success("计算完成！")

            if st.session_state.batch_results:
                res_df = pd.DataFrame(st.session_state.batch_results)
                cols_drop = ["尖峰消纳 (%)", "峰消纳 (%)", "平消纳 (%)", "谷消纳 (%)", "深谷消纳 (%)"]
                res_df = res_df.drop(columns=[c for c in cols_drop if c in res_df.columns])
                n_rows = len(res_df)
                table_height = (n_rows + 1) * 35 + 3
                st.dataframe(res_df.style.format({
                    "光伏容量 (MW)": "{:.2f}", "风电容量 (MW)": "{:.2f}",
                    "储能功率 (MW)": "{:.2f}", "储能时长 (h)": "{:.1f}",
                    "储能容量 (MWh)": "{:.2f}",
                    "加权自用电价": "{:.4f}", "加权上网电价": "{:.4f}", "综合电价": "{:.4f}",
                    "总发电量 (kWh)": "{:.2f}", "总消纳电量 (kWh)": "{:.2f}",
                    "总上网电量 (kWh)": "{:.2f}", "总折损电量 (kWh)": "{:.2f}",
                    "自用比例 (%)": "{:.2f}", "绿电占用电比例 (%)": "{:.2f}",
                    "光伏利用小时数 (h)": "{:.1f}", "风电利用小时数 (h)": "{:.1f}",
                    "储能等效循环次数": "{:.2f}"
                }), height=table_height, use_container_width=True)

                ex_data = write_batch_results_to_excel(st.session_state.batch_results, st.session_state.last_params)
                st.download_button(
                    label="下载 Excel 完整报表",
                    data=ex_data,
                    file_name=f"能源模拟分析_{datetime.datetime.now().strftime('%Y%m%d%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch'
                )


if __name__ == "__main__":
    st.set_page_config(layout="wide", page_title="新能源项目源荷储匹配分析平台")
    if check_login():
        main()
